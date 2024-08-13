import requests
from openpyxl import Workbook
from bs4 import BeautifulSoup
import time  # Импортируем библиотеку для работы с временем

# Функция для поиска текстового описания по запросу через Google
def get_description_from_google(query):
    search_url = 'https://www.google.com/search'
    params = {
        'q': query,
        'hl': 'en',  # Язык интерфейса
    }
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:90.0) Gecko/20100101 Firefox/90.0",
        "Accept-Language": "en-US,en;q=0.5",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive"
    }
    
    try:
        response = requests.get(search_url, params=params, headers=headers)
        response.raise_for_status()

        # Парсинг HTML ответа и поиск описаний
        soup = BeautifulSoup(response.text, 'lxml')
        
        # Ищем первый текстовый фрагмент, который может быть описанием
        description = None
        desc_tag = soup.find("span", {"class": "aCOpRe"})
        
        if desc_tag:
            description = desc_tag.get_text()

        return description
    except Exception as e:
        print(f"Ошибка при получении описания для {query}: {e}")
        return None

# Читаем данные из текстового файла
input_file_path = 'Тестовые_позиции.txt'
output_file_path = 'Тестовые_позиции_с_описанием.xlsx'

try:
    with open(input_file_path, 'r', encoding='utf-8') as file:
        content = file.read()
        print(f"Файл успешно прочитан.")
        items = [item.strip() for item in content.split(',') if item.strip()]
        print(f"Найдено {len(items)} артикулов.")
except Exception as e:
    print(f"Ошибка при чтении файла: {e}")
    items = []

# Создаем новый Excel файл
wb = Workbook()
ws = wb.active
ws.title = "Результаты"

# Добавляем заголовки
ws['A1'] = 'Артикул'
ws['B1'] = 'Описание'

# Проходим по каждому артикулу
for index, item_name in enumerate(items, start=2):
    if item_name:
        print(f"Ищем описание для: {item_name}")
        ws.cell(row=index, column=1, value=item_name)  # Записываем артикул в колонку A
        description = get_description_from_google(item_name)
        if description:
            ws.cell(row=index, column=2, value=description)  # Вставляем описание в колонку B
        else:
            print(f"Не удалось найти описание для: {item_name}")
    else:
        print(f"Пропускаем пустую строку: {index-1}")
    
    # Задержка перед следующим запросом
    time.sleep(3)  # Задержка в 3 секунды

# Сохраняем файл Excel
wb.save(output_file_path)

print(f"Результаты сохранены в файл {output_file_path}!")
