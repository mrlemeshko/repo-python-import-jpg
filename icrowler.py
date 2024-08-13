from icrawler.builtin import GoogleImageCrawler
from icrawler import ImageDownloader
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
import os
from io import BytesIO
from PIL import Image as PILImage
import shutil

# Папка для сохранения изображений
output_folder = 'downloaded_images'

# Чтение запросов из файла
def read_queries_from_file(filename):
    try:
        with open(filename, 'r', encoding='utf-8') as file:
            content = file.read()
            queries = [query.strip() for query in content.split(',') if query.strip()]
            print(f"Файл успешно прочитан. Найдено {len(queries)} запросов.")
            return queries
    except Exception as e:
        print(f"Ошибка при чтении файла: {e}")
        return []

# Очистка папки с изображениями
def clear_image_folder(folder):
    if os.path.exists(folder):
        shutil.rmtree(folder)
    os.makedirs(folder)

# Кастомный загрузчик для сохранения URL
class MyImageDownloader(ImageDownloader):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.downloaded_list = []

    def download(self, task, default_ext, timeout=5, max_retry=3, **kwargs):
        try:
            url = task['file_url']
            self.downloaded_list.append(url)
            # Передаем все аргументы в родительский метод
            return super().download(task, default_ext, timeout, max_retry, **kwargs)
        except Exception as e:
            print(f"Ошибка загрузки изображения: {e}")

# Функция для загрузки изображений с Google и вставки их в Excel
def download_images_and_insert_to_excel(queries, max_num=1):
    # Создаем новый Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты"

    # Добавляем заголовки
    ws['A1'] = 'Артикул'
    ws['B1'] = 'URL изображения'
    ws['C1'] = 'Изображение'

    for index, query in enumerate(queries, start=2):
        print(f"Ищем изображения для: {query}")
        
        # Очищаем папку перед каждым запросом
        clear_image_folder(output_folder)

        # Создаем экземпляр GoogleImageCrawler с кастомным загрузчиком
        google_crawler = GoogleImageCrawler(storage={'root_dir': output_folder},
                                            downloader_cls=MyImageDownloader)

        # Запуск поиска изображений
        google_crawler.crawl(keyword=query, max_num=max_num)

        # Проверяем, был ли найден URL изображения
        if google_crawler.downloader.downloaded_list:
            img_url = google_crawler.downloader.downloaded_list[0]  # Берем первый найденный URL
            img_path = os.path.join(output_folder, os.listdir(output_folder)[0])
            
            # Открываем изображение, адаптируем его и вставляем в Excel
            img = PILImage.open(img_path)
            img = img.resize((100, 100), PILImage.LANCZOS)  # Адаптируем изображение до 100x100
            buffer = BytesIO()
            img.save(buffer, format="JPEG")
            img_excel = ExcelImage(BytesIO(buffer.getvalue()))

            # Устанавливаем размеры строки и столбца для ячейки с изображением
            ws.row_dimensions[index].height = 75  # Высота строки
            ws.column_dimensions['C'].width = 15  # Ширина столбца

            # Записываем артикул и URL изображения
            ws.cell(row=index, column=1, value=query)
            ws.cell(row=index, column=2, value=img_url)  # Вставляем полный URL изображения
            img_excel.anchor = f'C{index}'
            ws.add_image(img_excel)
        else:
            print(f"Не удалось найти изображения для: {query}")

    # Сохраняем файл Excel
    wb.save('Тестовые_позиции_с_изображениями.xlsx')
    print("Результаты сохранены в файл 'Тестовые_позиции_с_изображениями.xlsx'")

# Основной код
if __name__ == "__main__":
    # Чтение запросов из файла
    queries = read_queries_from_file('Тестовые_позиции.txt')

    if queries:
        # Загрузка изображений и вставка в Excel
        download_images_and_insert_to_excel(queries)
